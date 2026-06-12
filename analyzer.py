#!/usr/bin/env python3
"""
Warehouse Auction Analyzer - Termux Version
Usage: python3 analyzer.py <excel_file.xlsx>
"""
import sys
import os
try:
    import openpyxl
except ImportError:
    os.system('pip install openpyxl')
    import openpyxl
from datetime import datetime

def pf(v):
    if v is None or v == '' or v == '#N/A': return 0
    if isinstance(v, (int, float)): return float(v) if v == v else 0  # nan check
    try: return float(str(v).replace(',',''))
    except: return 0

def fmt(v):
    v = round(v, 2)
    if v >= 100000: return f'₹{v/100000:.2f}L'
    if v >= 1000: return f'₹{v/1000:.1f}K'
    return f'₹{v:.2f}'

def main():
    if len(sys.argv) < 2:
        print("Usage: python3 analyzer.py <excel_file.xlsx> [warehouse_id]")
        print("Example: python3 analyzer.py Grocery_10_06_2026.xlsx")
        sys.exit(1)

    fname = sys.argv[1]
    if not os.path.exists(fname):
        print(f"Error: File '{fname}' not found!")
        sys.exit(1)

    print(f"\n📦 Warehouse Auction Analyzer")
    print(f"📁 File: {fname}")
    print(f"{'='*60}")

    wb = openpyxl.load_workbook(fname, data_only=True)
    ws = wb[wb.SheetNames[0]]
    headers = [cell.value for cell in ws[1]]

    # Find columns
    col = {}
    for i, h in enumerate(headers):
        hl = str(h).lower() if h else ''
        if 'warehouse' in hl: col['wh'] = i
        elif 'brand' in hl: col['brand'] = i
        elif 'product' in hl: col['title'] = i
        elif hl == 'mrp': col['mrp'] = i
        elif 'quantity' in hl or 'qty' in hl: col['qty'] = i
        elif 'mfg' in hl or 'manufacturing' in hl: col['mfg'] = i
        elif 'expiry' in hl or 'expire' in hl: col['exp'] = i
        elif 'aging' in hl or hl == 'age': col['aging'] = i
        elif 'slab' in hl: col['slab'] = i
        elif 'lot' in hl: col['lot'] = i
        elif 'wh type' in hl: col['whtype'] = i
        elif 'category' in hl: col['cat'] = i
        elif 'yield' in hl: col['yield'] = i
        elif 'tenative' in hl or 'tentative' in hl: col['tenative'] = i
        elif 'tcs value' in hl: col['tcsval'] = i
        elif hl == 'tcs': col['tcs'] = i
        elif 'tentative final' in hl: col['tf'] = i
        elif 'final mrp' in hl: col['fm'] = i

    if 'wh' not in col:
        print("Error: warehouse_id column not found!")
        sys.exit(1)

    # Get all rows
    all_rows = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if any(c is not None and c != '' for c in row):
            all_rows.append(row)

    # Get unique warehouses
    whs = sorted(set(r[col['wh']] for r in all_rows if r[col['wh']]))
    
    if not whs:
        print("Error: No warehouse data found!")
        sys.exit(1)

    # Select warehouse
    target_wh = None
    if len(sys.argv) >= 3:
        target_wh = sys.argv[2]
        if target_wh not in whs:
            print(f"Error: Warehouse '{target_wh}' not found!")
            print(f"Available: {', '.join(whs)}")
            sys.exit(1)
    elif len(whs) == 1:
        target_wh = whs[0]
    else:
        print(f"\n🏭 Available Warehouses ({len(whs)}):")
        for i, w in enumerate(whs, 1):
            count = sum(1 for r in all_rows if r[col['wh']] == w)
            print(f"  [{i}] {w} ({count} products)")
        print(f"\nEnter number (1-{len(whs)}): ", end='')
        try:
            choice = int(input().strip())
            target_wh = whs[choice - 1]
        except:
            print("Invalid choice!")
            sys.exit(1)

    print(f"\n🏭 Analyzing: {target_wh}")
    print(f"{'='*60}")

    # Filter rows
    rows = [r for r in all_rows if r[col['wh']] == target_wh]
    today = datetime.now()
    today = today.replace(hour=0, minute=0, second=0, microsecond=0)

    prods = []
    tQ = tTF = tFM = tMRP = 0

    for r in rows:
        mrp = pf(r[col.get('mrp', 12)])
        qty = pf(r[col.get('qty', 13)])
        tf = pf(r[col.get('tf', 28)])
        fm = pf(r[col.get('fm', 29)])
        aging = pf(r[col.get('aging', 21)])
        yield_pct = pf(r[col.get('yield', 24)])
        tenative = pf(r[col.get('tenative', 25)])
        tcs = pf(r[col.get('tcs', 26)])
        tcs_val = pf(r[col.get('tcsval', 27)])

        days_left = aging
        if r[col.get('exp', 17)]:
            try:
                ed = r[col['exp']]
                if isinstance(ed, datetime):
                    days_left = (ed - today).days
            except: pass

        if days_left <= 30: pri = '🔴 CRITICAL'
        elif days_left <= 60: pri = '🟡 MEDIUM'
        else: pri = '🟢 SAFE'

        mfg_s = ''
        if r[col.get('mfg', 16)]:
            try:
                mfg_s = r[col['mfg']].strftime('%d-%b-%Y') if isinstance(r[col['mfg']], datetime) else str(r[col['mfg']])[:10]
            except: mfg_s = '-'

        exp_s = ''
        if r[col.get('exp', 17)]:
            try:
                exp_s = r[col['exp']].strftime('%d-%b-%Y') if isinstance(r[col['exp']], datetime) else str(r[col['exp']])[:10]
            except: exp_s = '-'

        prods.append({
            'brand': r[col.get('brand', 10)] or '',
            'title': r[col.get('title', 11)] or '',
            'cat': r[col.get('cat', 8)] or '',
            'mrp': mrp, 'qty': qty, 'mfg': mfg_s, 'exp': exp_s,
            'days': int(days_left), 'slab': r[col.get('slab', 22)] or '',
            'yield': yield_pct, 'tenative': tenative, 'tcs': tcs,
            'tcs_val': tcs_val, 'tf': tf, 'fm': fm, 'pri': pri
        })
        tQ += qty; tTF += tf; tFM += fm; tMRP += mrp * qty

    disc = ((tMRP - tTF) / tMRP * 100) if tMRP > 0 else 0

    # Print results
    print(f"\n📊 RESULTS")
    print(f"{'─'*60}")
    print(f"  Products:         {len(prods)}")
    print(f"  Total Qty:        {int(tQ):,}")
    print(f"  MRP Value:        {fmt(tMRP)}")
    print(f"  Tentative Final:  {fmt(tTF)}")
    print(f"  Final MRP:        {fmt(tFM)}")
    print(f"  Discount:         {disc:.1f}%")
    print(f"{'─'*60}")

    if disc >= 25: print(f"\n  ✅ STRONG BUY — High discount, great deal!")
    elif disc >= 15: print(f"\n  🟡 CONSIDER — Moderate discount, negotiate")
    elif disc >= 10: print(f"\n  ⚠️ WEAK BUY — Low discount, try to negotiate")
    else: print(f"\n  ❌ SKIP — Not enough discount")

    # Priority breakdown
    crit = [p for p in prods if 'CRITICAL' in p['pri']]
    med = [p for p in prods if 'MEDIUM' in p['pri']]
    safe = [p for p in prods if 'SAFE' in p['pri']]

    print(f"\n📦 Priority Breakdown:")
    print(f"  🔴 Critical (≤30d): {len(crit)} products, TF: {fmt(sum(p['tf'] for p in crit))}")
    print(f"  🟡 Medium (31-60d): {len(med)} products, TF: {fmt(sum(p['tf'] for p in med))}")
    print(f"  🟢 Safe (60+d):     {len(safe)} products, TF: {fmt(sum(p['tf'] for p in safe))}")

    # Product table
    print(f"\n{'─'*100}")
    print(f"  {'#':<4} {'Brand':<16} {'Product':<35} {'MRP':>8} {'Qty':>6} {'Expiry':<12} {'Days':>5} {'Tentative Final':>16} {'Pri':<10}")
    print(f"{'─'*100}")
    for i, p in enumerate(prods, 1):
        title = p['title'][:33] if p['title'] else ''
        brand = p['brand'][:14] if p['brand'] else ''
        print(f"  {i:<4} {brand:<16} {title:<35} ₹{p['mrp']:>6.0f} {int(p['qty']):>6} {p['exp']:<12} {p['days']:>5} ₹{p['tf']:>14.2f} {p['pri']:<10}")
    print(f"{'─'*100}")
    print(f"  {'TOTAL':<57} {int(tQ):>6} {'':>18} ₹{tTF:>14.2f}")
    print(f"{'─'*100}")

    # Critical items
    if crit:
        print(f"\n🔴 CRITICAL ITEMS (≤30 days to expire):")
        crit_sorted = sorted(crit, key=lambda x: x['days'])
        for i, p in enumerate(crit_sorted, 1):
            print(f"  [{i}] {p['brand']} - {p['title'][:40]}")
            print(f"      Qty: {int(p['qty'])} | MRP: ₹{p['mrp']:.0f} | Days: {p['days']} | TF: ₹{p['tf']:.2f}")

    # Category breakdown
    cats = {}
    for p in prods:
        c = p['cat'] or 'Unknown'
        cats[c] = cats.get(c, 0) + p['qty']
    print(f"\n📦 Categories:")
    for c, q in sorted(cats.items(), key=lambda x: -x[1]):
        pct = (q / tQ * 100) if tQ > 0 else 0
        print(f"  {c:<25} {int(q):>8,} units ({pct:.1f}%)")

    # Generate Excel report
    print(f"\n📥 Generating Excel report...")
    out_wb = openpyxl.Workbook()
    
    # Sheet 1: Analysis
    ws1 = out_wb.active
    ws1.title = 'Analysis'
    h = ['S.No','Brand','Product','Category','MRP','Qty','Mfg','Expiry','Days','Slab','Yield','Tenative','TCS','TCS Val','Tentative Final','Final MRP','Priority']
    for ci, hv in enumerate(h, 1):
        c = ws1.cell(row=1, column=ci, value=hv)
        c.font = openpyxl.styles.Font(bold=True, color='FFFFFF')
        c.fill = openpyxl.styles.PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    for ri, p in enumerate(prods, 2):
        vals = [ri-1, p['brand'], p['title'], p['cat'], p['mrp'], int(p['qty']), p['mfg'], p['exp'], p['days'], p['slab'], p['yield'], p['tenative'], p['tcs'], p['tcs_val'], p['tf'], p['fm'], p['pri']]
        for ci, v in enumerate(vals, 1):
            ws1.cell(row=ri, column=ci, value=v)
    tr = len(prods) + 2
    ws1.cell(row=tr, column=1, value='TOTAL').font = openpyxl.styles.Font(bold=True)
    ws1.cell(row=tr, column=5, value=int(tQ)).font = openpyxl.styles.Font(bold=True)
    ws1.cell(row=tr, column=15, value=round(tTF, 2)).font = openpyxl.styles.Font(bold=True)
    ws1.cell(row=tr, column=16, value=round(tFM, 2)).font = openpyxl.styles.Font(bold=True)

    # Sheet 2: Decision
    ws2 = out_wb.create_sheet('Decision')
    ws2.append(['PURCHASE DECISION'])
    ws2.append([])
    ws2.append(['Products', len(prods)])
    ws2.append(['Quantity', int(tQ)])
    ws2.append(['MRP Value', f'₹{tMRP:.0f}'])
    ws2.append(['Tentative Final', f'₹{tTF:.2f}'])
    ws2.append(['Final MRP', f'₹{tFM:.2f}'])
    ws2.append(['Discount %', f'{disc:.1f}%'])
    ws2.append([])
    ws2.append(['Verdict', 'STRONG BUY' if disc >= 25 else 'CONSIDER' if disc >= 15 else 'WEAK BUY' if disc >= 10 else 'SKIP'])
    ws2.append([])
    ws2.append(['Priority', 'Products', 'Tentative Final', 'Final MRP'])
    ws2.append(['CRITICAL', len(crit), f'₹{sum(p["tf"] for p in crit):.2f}', f'₹{sum(p["fm"] for p in crit):.2f}'])
    ws2.append(['MEDIUM', len(med), f'₹{sum(p["tf"] for p in med):.2f}', f'₹{sum(p["fm"] for p in med):.2f}'])
    ws2.append(['SAFE', len(safe), f'₹{sum(p["tf"] for p in safe):.2f}', f'₹{sum(p["fm"] for p in safe):.2f}'])

    # Sheet 3: Critical
    ws3 = out_wb.create_sheet('Critical')
    ws3.append(['#','Brand','Product','MRP','Qty','Expiry','Days Left','Tentative Final','Final MRP'])
    for ri, p in enumerate(sorted(crit, key=lambda x: x['days']), 1):
        ws3.append([ri, p['brand'], p['title'], p['mrp'], int(p['qty']), p['exp'], p['days'], p['tf'], p['fm']])

    out_fname = f'Auction_{target_wh.replace(" ","_")}.xlsx'
    out_wb.save(out_fname)
    print(f"✅ Report saved: {out_fname}")
    print(f"\n💡 Tips:")
    print(f"  • Critical items mein zyada discount mil sakta hai")
    print(f"  • Bid Tentative Final se 10-15% kam pe karo")
    print(f"  • High MRP items pe focus karo")

if __name__ == '__main__':
    main()
